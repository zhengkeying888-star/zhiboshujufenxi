"""
排期Excel解析导入脚本
==================
功能：
1. 解析《平台私域直播线宣发排期【4-5月】.xlsx》
2. 将二维排期表扁平化为一维记录
3. 导入到飞书排期明细多维表

使用方法：
    1. 先运行 setup_schedule_bitable.py 创建排期多维表
    2. 将 SCHEDULE_APP_TOKEN / SCHEDULE_TABLE_ID 填入 config_local.py
    3. python import_schedule_excel.py
"""

import sys
import os
import re
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from config_local import APP_ID, APP_SECRET, SCHEDULE_APP_TOKEN, SCHEDULE_TABLE_ID
except ImportError:
    from config import APP_ID, APP_SECRET, SCHEDULE_APP_TOKEN, SCHEDULE_TABLE_ID

from feishu_client import FeishuClient

# 排期Excel路径
SCHEDULE_EXCEL = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "平台私域直播线宣发排期【4-5月】.xlsx"
)

# ========== 品类相关（复用 deep_dive_analysis.py） ==========
STANDARD_CATS = [
    "健康营养", "太极", "五禽戏", "睡眠调理", "气血调理", "固气活血", "君合太极",
    "开心太极", "内养太极", "云帆太极", "东方食养", "古法居家养生", "华佗肩颈舒活功",
    "健康家厨", "健康食养", "儿童健康", "食养助长", "体质食养", "易筋经", "营养调理",
    "中式美食制作", "轻训营", "亚健康管理", "私域",
    "普拉提", "瑜伽", "中医变美", "穿搭", "懒人吃瘦", "面部瑜伽驻颜", "逆龄女神瑜伽",
    "逆龄普拉提", "女性保养瑜伽", "东方养正瑜伽", "塑形流瑜伽", "体态", "体态塑形瑜伽",
    "形体芭蕾", "养正变美", "一杰瑜伽", "正位塑形瑜伽", "瑜伽会员",
    "手机摄影", "摄影美学", "唱歌", "短视频", "风光摄影", "相机摄影", "声乐", "国际声乐",
    "电子琴", "键盘乐", "真书法", "油画", "国画", "国学朗诵", "戏曲", "舞蹈", "优雅舞蹈",
    "茶道", "编织工艺美学", "钩针编织美学", "美学收纳",
    "中医瑜伽", "面部驻颜瑜伽",
]

CAT_ALIASES = {
    "居家古法": "古法居家养生",
    "气血": "气血调理",
    "睡眠": "睡眠调理",
    "五禽戏晨练": "五禽戏",
    "普拉提晨练": "普拉提",
    "瑜伽晨练": "瑜伽",
    "君合太极晨练": "君合太极",
    "逆龄女神瑜伽晨练": "逆龄女神瑜伽",
    "东方养正瑜伽晨练": "东方养正瑜伽",
    "一杰晨练": "一杰瑜伽",
    "晨练焕醒计划": "晨练",
    "节气养正公开课": "节气养正",
    "老师，请回答！": "健康营养",
    "才艺名师公开课": "才艺",
    "元气干货分享课": "元气干货",
    "中医瑜伽-陈浙南": "中医瑜伽",
    "瑜伽：杨淇": "瑜伽",
    "固气": "固气活血",
    "亚健康": "亚健康管理",
    # 补充解析中的未识别项
    "姚国诚单人": "古法居家养生",
    "古法居家姚国诚": "古法居家养生",
    "4.9姚国诚单人": "古法居家养生",
    "陈浙南0元新体验营": "中医瑜伽",
    "王溪0元（新栏目）": "中医变美",
    "写作课": "健康营养",  # 需要确认
    "朗诵IP": "国学朗诵",
    "编织-0元": "编织工艺美学",
    "影像一点通": "摄影美学",
    "魏巍异地": "普拉提",  # 魏巍是普拉提老师
    "收纳，不需要剪辑": "美学收纳",
    "2026.4.2唐一杰": "一杰瑜伽",
}


def normalize_category(name):
    """品类规范化"""
    if not name:
        return None
    name = str(name).strip()
    # 1. 精确匹配
    if name in STANDARD_CATS:
        return name
    # 2. 别名匹配
    if name in CAT_ALIASES:
        return CAT_ALIASES[name]
    # 3. 去除常见后缀
    for suffix in ["晨练", "带练", "公开课", "直播", "IP", "大赛"]:
        if name.endswith(suffix):
            base = name[: -len(suffix)].strip()
            if base in STANDARD_CATS:
                return base
            if base in CAT_ALIASES:
                return CAT_ALIASES[base]
    # 4. 子串包含匹配（优先最长）
    for cat in sorted(STANDARD_CATS, key=len, reverse=True):
        if cat in name:
            return cat
    for alias in sorted(CAT_ALIASES.keys(), key=len, reverse=True):
        if alias in name:
            return CAT_ALIASES[alias]
    return None


# ========== Excel解析 ==========

def calc_formula(val):
    """计算简单Excel公式（仅支持加法）"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s.startswith("="):
        nums = re.findall(r"\d+", s)
        return sum(int(n) for n in nums)
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def parse_date_cell(val, default_month=None, default_year=2026):
    """解析日期单元格，返回 (month, day)"""
    if val is None:
        return None, None
    s = str(val).strip()
    # 4.6, 4.7 格式
    m = re.match(r"(\d+)\.(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    # 纯数字
    try:
        day = int(float(s))
        if 1 <= day <= 31:
            return default_month, day
    except ValueError:
        pass
    return None, None


def infer_month_from_sheet_name(sheet_name):
    """从sheet名推断月份"""
    # 4.6-4.12 -> 4月
    # 5.11-5.17 -> 5月
    m = re.search(r"(\d+)月", sheet_name)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\.\d+", sheet_name)
    if m:
        return int(m.group(1))
    # 4月排期 -> 4
    m = re.search(r"(\d+)月排期", sheet_name)
    if m:
        return int(m.group(1))
    # 4.13-4.19 取第一个数字作为月
    m = re.search(r"(\d+)\.\d+", sheet_name)
    if m:
        return int(m.group(1))
    return None


def extract_live_blocks(cell_text):
    """
    从单元格文本提取直播块列表。
    每个块是一个 dict: {name_lines, time, tags}
    """
    if not cell_text:
        return []
    raw_lines = str(cell_text).split('\n')
    lines = [l.strip() for l in raw_lines if l.strip()]
    if not lines:
        return []

    blocks = []
    current_block = {"lines": [], "time": "", "tags": []}

    # 标记关键词
    tag_keywords = ["复用", "需剪辑", "数字人", "录播", "不回捞", "已有单课id", "伪直播"]
    time_pattern = re.compile(r"\d{1,2}[:：]\d{2}(-\d{1,2}[:：]\d{2})?")
    url_pattern = re.compile(r"https?://")

    for line in lines:
        # 跳过URL
        if url_pattern.search(line):
            continue
        # 跳过纯预约链接提示
        if line.startswith("预约链接：") or line.startswith("开播时间："):
            continue
        # 识别时间
        if time_pattern.match(line):
            current_block["time"] = line.replace("：", ":")
            continue
        # 识别标记
        is_tag = False
        for tag in tag_keywords:
            if tag in line:
                current_block["tags"].append(tag)
                is_tag = True
                break
        if is_tag:
            continue
        # 如果这一行明显是新直播名（包含品类关键词），且当前块已有内容，则开新块
        cat = normalize_category(line)
        if cat and current_block["lines"]:
            # 检查是否前一个也是同类（如"一杰瑜伽晨练"后面跟"五禽戏晨练"）
            blocks.append(current_block)
            current_block = {"lines": [line], "time": "", "tags": []}
        else:
            current_block["lines"].append(line)

    if current_block["lines"]:
        blocks.append(current_block)

    return blocks


def parse_schedule_excel(filepath):
    """解析排期Excel，返回一维记录列表"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=False)
    records = []
    stats = defaultdict(int)
    unmapped = set()

    for sheet_name in wb.sheetnames:
        if "排期" not in sheet_name or "人数" in sheet_name:
            continue

        ws = wb[sheet_name]
        sheet_month = infer_month_from_sheet_name(sheet_name)

        # 收集所有行内容用于定位（扫描前3列）
        rows_content = []
        for i in range(1, min(ws.max_row + 1, 30)):
            row_vals = [ws.cell(i, j).value for j in range(1, min(ws.max_column + 1, 12))]
            rows_content.append((i, row_vals))

        # 找日期行：包含"日期"文字或连续数字的行
        date_row_idx = None
        date_cols = []  # [(col_idx, month, day)]

        for i, row_vals in rows_content:
            row_str = " ".join(str(v) for v in row_vals if v is not None)
            if "日期" in row_str:
                date_row_idx = i
                # 日期通常在下一行
                if i + 1 <= ws.max_row:
                    date_row_vals = [ws.cell(i + 1, j).value for j in range(1, min(ws.max_column + 1, 12))]
                else:
                    date_row_vals = row_vals
                for j, val in enumerate(date_row_vals, 1):
                    m, d = parse_date_cell(val, default_month=sheet_month)
                    if m and d:
                        date_cols.append((j, m, d))
                if not date_cols:
                    # 可能日期就在当前行
                    for j, val in enumerate(row_vals, 1):
                        m, d = parse_date_cell(val, default_month=sheet_month)
                        if m and d:
                            date_cols.append((j, m, d))
                break
            # 尝试直接匹配日期数字行
            date_nums = []
            for j, val in enumerate(row_vals, 1):
                m, d = parse_date_cell(val, default_month=sheet_month)
                if m and d:
                    date_nums.append((j, m, d))
            if len(date_nums) >= 3:
                date_row_idx = i
                date_cols = date_nums
                break

        if not date_cols:
            print(f"   ⚠️ {sheet_name}: 未找到日期行，跳过")
            continue

        # 辅助函数：在前3列中查找关键词
        def row_has_keyword(row_vals, keyword):
            for v in row_vals[:3]:
                if v and keyword in str(v):
                    return True
            return False

        # 找时段行：早间、晚IP专场、晚上平播、伪直播复用
        slot_rows = []  # [(row_idx, slot_type)]
        exposure_row_map = {}  # row_idx -> 曝光行idx
        writer_row_map = {}  # row_idx -> 文案负责人行idx

        for i, row_vals in rows_content:
            first_col = str(row_vals[0]) if row_vals and row_vals[0] else ""
            if first_col in ["早间", "早上"]:
                slot_rows.append((i, "晨练"))
            elif first_col in ["晚IP专场", "晚上平播", "晚间"]:
                slot_rows.append((i, "晚间"))
            elif first_col in ["伪直播复用"]:
                slot_rows.append((i, "伪直播"))
            elif row_has_keyword(row_vals, "曝光量级"):
                # 记录这是哪个时段的曝光行
                for prev_i, slot_type in slot_rows:
                    if i > prev_i and i < prev_i + 5:
                        exposure_row_map[prev_i] = i
                        break
            elif row_has_keyword(row_vals, "文案负责人"):
                for prev_i, slot_type in slot_rows:
                    if i > prev_i and i < prev_i + 5:
                        writer_row_map[prev_i] = i
                        break

        if not slot_rows:
            print(f"   ⚠️ {sheet_name}: 未找到时段行，跳过")
            continue

        print(f"   📖 {sheet_name}: 找到 {len(date_cols)} 个日期, {len(slot_rows)} 个时段")

        for slot_row_idx, slot_type in slot_rows:
            # 找到对应的曝光行和文案负责人行
            exposure_row_idx = exposure_row_map.get(slot_row_idx)
            writer_row_idx = writer_row_map.get(slot_row_idx)

            for col_idx, month, day in date_cols:
                cell_val = ws.cell(slot_row_idx, col_idx).value
                if not cell_val:
                    continue

                # 获取曝光
                exposure = 0
                if exposure_row_idx:
                    exp_val = ws.cell(exposure_row_idx, col_idx).value
                    exposure = calc_formula(exp_val)

                # 获取文案负责人
                writer = ""
                if writer_row_idx:
                    w_val = ws.cell(writer_row_idx, col_idx).value
                    if w_val:
                        writer = str(w_val).strip()

                # 解析直播块
                blocks = extract_live_blocks(cell_val)
                for block in blocks:
                    name = block["lines"][0] if block["lines"] else ""
                    cat = normalize_category(name)

                    if not cat:
                        unmapped.add(name)
                        stats["unmapped"] += 1
                        continue

                    # 判断线级
                    line_level = ""
                    health_cats = ["五禽戏", "君合太极", "开心太极", "内养太极", "云帆太极", "太极", "太极s", "太极A", "太极BCD",
                                   "睡眠调理", "气血调理", "固气活血", "健康营养", "健康食养", "东方食养", "体质食养", "易筋经",
                                   "营养调理", "中式美食制作", "轻训营", "亚健康管理", "亚健康", "儿童健康", "健康家厨",
                                   "食养助长", "华佗肩颈舒活功", "古法居家养生", "私域"]
                    beauty_cats = ["瑜伽", "瑜伽S", "瑜伽A", "瑜伽BCD", "瑜伽会员", "普拉提", "普拉提S", "普拉提A", "普拉提BCD",
                                   "逆龄女神瑜伽", "体态塑形瑜伽", "正位塑形瑜伽", "一杰瑜伽", "东方养正瑜伽", "塑形流瑜伽",
                                   "女性保养瑜伽", "面部瑜伽驻颜", "懒人吃瘦", "养正变美", "中医变美", "中医瑜伽", "美学",
                                   "形体芭蕾", "体态", "穿搭", "面部驻颜瑜伽"]
                    interest_cats = ["唱歌", "声乐", "国际声乐", "短视频", "摄影美学", "手机摄影", "手机摄影BCD", "相机摄影",
                                     "风光摄影", "舞蹈", "优雅舞蹈", "戏曲", "真书法", "油画", "国画", "国学朗诵", "茶道",
                                     "编织工艺美学", "钩针编织美学", "美学收纳", "电子琴", "键盘乐"]
                    if cat in health_cats:
                        line_level = "健康线"
                    elif cat in beauty_cats:
                        line_level = "变美线"
                    elif cat in interest_cats:
                        line_level = "兴趣线"

                    date_obj = datetime(2026, month, day)
                    month_label = f"{month}月"

                    # 构建直播名（除第一行外的其他描述）
                    live_name = " ".join(block["lines"]) if len(block["lines"]) > 1 else name

                    records.append({
                        "日期": date_obj,
                        "月份": month_label,
                        "品类": cat,
                        "直播名": live_name,
                        "时段": slot_type,
                        "曝光量级": exposure,
                        "标记": ",".join(block["tags"]) if block["tags"] else "",
                        "线级": line_level,
                        "文案负责人": writer,
                        "所属周次": sheet_name,
                        "时间": block["time"],
                    })
                    stats["mapped"] += 1

    print(f"\n📊 解析统计:")
    print(f"   成功识别: {stats['mapped']} 条")
    print(f"   未识别品类: {stats['unmapped']} 条")
    if unmapped:
        print(f"   未识别项: {', '.join(list(unmapped)[:20])}")

    return records


def import_to_bitable(records):
    """导入到飞书多维表"""
    if not SCHEDULE_APP_TOKEN or not SCHEDULE_TABLE_ID:
        print("\n❌ 错误：请先在 config_local.py 中配置 SCHEDULE_APP_TOKEN 和 SCHEDULE_TABLE_ID")
        print("   运行: python setup_schedule_bitable.py")
        sys.exit(1)

    client = FeishuClient(APP_ID, APP_SECRET)

    # 准备飞书记录格式
    feishu_records = []
    for r in records:
        record = {
            "日期": int(r["日期"].timestamp() * 1000),
            "月份": r["月份"],
            "品类": r["品类"],
            "直播名": r["直播名"],
            "时段": r["时段"],
            "曝光量级": r["曝光量级"],
            "标记": r["标记"],
            "线级": r["线级"],
            "文案负责人": r["文案负责人"],
            "所属周次": r["所属周次"],
            "时间": r["时间"],
        }
        feishu_records.append(record)

    print(f"\n📤 导入到飞书多维表 ...")
    created_ids = client.batch_create_records(SCHEDULE_APP_TOKEN, SCHEDULE_TABLE_ID, feishu_records)
    print(f"✅ 成功导入 {len(created_ids)} 条记录")
    return created_ids


def main():
    print("=" * 50)
    print("排期Excel解析导入")
    print("=" * 50)

    if not os.path.exists(SCHEDULE_EXCEL):
        print(f"\n❌ 错误：排期Excel未找到: {SCHEDULE_EXCEL}")
        sys.exit(1)

    print(f"\n📖 解析排期Excel: {os.path.basename(SCHEDULE_EXCEL)}")
    records = parse_schedule_excel(SCHEDULE_EXCEL)

    if not records:
        print("\n❌ 未解析到任何排期记录")
        sys.exit(1)

    # 按日期排序
    records.sort(key=lambda x: x["日期"])

    print(f"\n📋 预览前10条:")
    for r in records[:10]:
        print(f"   {r['日期'].strftime('%m-%d')} | {r['时段']} | {r['品类']} | {r['直播名'][:30]} | 曝光:{r['曝光量级']}")

    # 导入
    import_to_bitable(records)

    print("\n" + "=" * 50)
    print("🎉 导入完成！")
    print("=" * 50)


if __name__ == "__main__":
    main()
